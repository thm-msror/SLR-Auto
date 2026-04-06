from __future__ import annotations

import json
import sys
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_LOG_PATH = Path(__file__).with_name("example.json")
DEFAULT_TEST_OUTPUT_PATH = DEFAULT_LOG_PATH.with_suffix(".test.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUBTLE_FILL = PatternFill(fill_type="solid", fgColor="F7F9FC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def load_log(log_path: str | Path) -> dict[str, Any]:
    path = Path(log_path)
    return json.loads(path.read_text(encoding="utf-8"))


def export_log_to_excel(log_path: str | Path, output_path: str | Path) -> Path:
    run = load_log(log_path)
    return export_run_to_excel(run, output_path)


def export_run_to_excel(run: dict[str, Any], output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = _build_workbook(run)
    wb.save(output)
    return output


def export_run_to_excel_bytes(run: dict[str, Any]) -> bytes:
    wb = _build_workbook(run)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_workbook(run: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _build_user_input_sheet(wb, run)
    _build_initial_screen_sheet(wb, run)
    _build_final_paper_reading_sheet(wb, run)
    return wb


def _build_user_input_sheet(wb: Workbook, run: dict[str, Any]) -> None:
    ws = wb.create_sheet("User Input")
    inputs = run.get("inputs") or {}
    _write_section_title(ws, 1, "User Input Summary")
    criteria_specs = _build_criterion_specs(inputs.get("criteria_used") or [])
    top_rows = [
        [
            "Research Questions",
            "",
            _stringify(inputs.get("research_questions")),
        ],
        [
            "Boolean Query",
            _stringify(inputs.get("boolean_query_suggested")),
            _stringify(inputs.get("boolean_query_used")),
        ],
        [
            "Screening Criteria",
            _join_lines(inputs.get("criteria_suggested") or []),
            _join_lines(inputs.get("criteria_used") or []),
        ],
        [
            "Research Themes",
            _stringify(inputs.get("research_themes_suggested")) or _format_theme_text(run.get("categories") or {}),
            _stringify(inputs.get("research_themes_used")) or _format_theme_text(run.get("categories") or {}),
        ],
    ]
    top_end = _write_table(
        ws,
        start_row=2,
        start_col=1,
        headers=["Field", "Suggested", "Used"],
        rows=top_rows,
        table_name="UserInputTopTable",
    )
    ws.row_dimensions[3].height = 88
    ws.row_dimensions[4].height = 88
    ws.row_dimensions[5].height = 140
    ws.row_dimensions[6].height = 140

    criteria_start = top_end + 1
    current_row = _write_used_criteria_table(
        ws,
        start_row=criteria_start,
        criterion_specs=criteria_specs,
    )

    _set_column_widths(ws, {"A": 22, "B": 70, "C": 70})
    _set_row_heights(ws, criteria_start + 1, current_row, 30)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _build_initial_screen_sheet(wb: Workbook, run: dict[str, Any]) -> None:
    ws = wb.create_sheet("Initial Screen")
    inputs = run.get("inputs") or {}
    criteria_used = [str(item).strip() for item in (inputs.get("criteria_used") or []) if str(item).strip()]
    criterion_specs = _build_criterion_specs(criteria_used)
    headers = _initial_screen_headers(criterion_specs)
    rows = _initial_screen_rows(run, criterion_specs)

    _write_section_title(ws, 1, "Initial Screening Decisions")
    table_start = 2
    table_end = _write_table(
        ws,
        start_row=table_start,
        start_col=1,
        headers=headers,
        rows=rows,
        table_name="InitialScreenTable",
    )

    width_map = {
        "A": 28,
        "B": 44,
        "C": 22,
        "D": 14,
        "E": 28,
        "F": 34,
        "G": 14,
        "H": 12,
        "I": 12,
        "J": 16,
        "K": 12,
        "L": 12,
        "M": 16,
    }
    criterion_start_col = 14
    for idx in range(len(criterion_specs)):
        letter = get_column_letter(criterion_start_col + idx)
        width_map[letter] = 26
    _set_column_widths(ws, width_map)
    _set_row_heights(ws, table_start + 1, table_end, 32)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _build_final_paper_reading_sheet(wb: Workbook, run: dict[str, Any]) -> None:
    ws = wb.create_sheet("Final Paper Reading")
    categories = list((run.get("categories") or {}).keys())
    headers = ["Paper ID", "Title", "Included", "Publisher", "Published", "DOI", "URL", *categories]
    rows = _final_paper_reading_rows(run, categories)

    _write_section_title(ws, 1, "Full-Text Extraction by Category")
    table_start = 2
    table_end = _write_table(
        ws,
        start_row=table_start,
        start_col=1,
        headers=headers,
        rows=rows,
        table_name="FinalPaperReadingTable",
    )

    width_map = {
        "A": 28,
        "B": 44,
        "C": 12,
        "D": 22,
        "E": 14,
        "F": 28,
        "G": 34,
    }
    for idx in range(len(categories)):
        letter = get_column_letter(8 + idx)
        width_map[letter] = 54
    _set_column_widths(ws, width_map)
    _set_row_heights(ws, table_start + 1, table_end, 120)

    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False


def _initial_screen_headers(criterion_specs: list[dict[str, str]]) -> list[str]:
    return [
        "Paper ID",
        "Title",
        "Publisher",
        "Published",
        "DOI",
        "URL",
        "Relevance Score",
        "Include Yes",
        "Include No",
        "Include Insufficient",
        "Exclude Yes",
        "Exclude No",
        "Exclude Insufficient",
        *[_criterion_header_label(spec) for spec in criterion_specs],
    ]


def _initial_screen_rows(run: dict[str, Any], criterion_specs: list[dict[str, str]]) -> list[list[str | int | None]]:
    papers_by_id = run.get("papers_by_id") or {}
    sortable_rows: list[tuple[int, str, list[str | int | None]]] = []

    for paper_id, paper in papers_by_id.items():
        screening = paper.get("screening") or {}
        counts = screening.get("counts") or {}
        include_counts = counts.get("include") or {}
        exclude_counts = counts.get("exclude") or {}
        answers = screening.get("answers") or []
        answer_map = {
            str(item.get("criterion") or "").strip(): str(item.get("answer") or "").strip()
            for item in answers
            if item
        }
        ordered_answers = [
            _normalize_screening_answer(item.get("answer"))
            for item in answers
            if item
        ]

        link = paper.get("link")
        if not link and paper.get("doi"):
            link = f"https://doi.org/{paper.get('doi')}"

        relevance_score = screening.get("relevance_score")
        sort_score = relevance_score if isinstance(relevance_score, (int, float)) else -1
        title = str(paper.get("title") or "")
        row: list[str | int | None] = [
            paper_id,
            title,
            paper.get("publisher") or "",
            paper.get("published") or "",
            paper.get("doi") or "",
            link or "",
            relevance_score,
            include_counts.get("yes"),
            include_counts.get("no"),
            include_counts.get("insufficient"),
            exclude_counts.get("yes"),
            exclude_counts.get("no"),
            exclude_counts.get("insufficient"),
        ]
        for idx, spec in enumerate(criterion_specs):
            answer = _normalize_screening_answer(answer_map.get(spec["criterion"]))
            if not answer and idx < len(ordered_answers):
                answer = ordered_answers[idx]
            row.append(answer or "INSUFFICIENT")
        sortable_rows.append((int(sort_score), title.lower(), row))

    sortable_rows.sort(key=lambda item: (-item[0], item[1]))
    return [row for _, _, row in sortable_rows]


def _final_paper_reading_rows(run: dict[str, Any], categories: list[str]) -> list[list[str]]:
    papers_by_id = run.get("papers_by_id") or {}
    top_papers = run.get("top_paper_ids") or {}
    sortable_rows: list[tuple[int, str, list[str]]] = []

    for paper_id, entry in top_papers.items():
        paper = papers_by_id.get(paper_id) or {}
        full_screening = entry.get("full_screening") or {}
        category_map = full_screening.get("categories") or {}
        included = full_screening.get("included")

        link = paper.get("link")
        if not link and paper.get("doi"):
            link = f"https://doi.org/{paper.get('doi')}"

        title = str(entry.get("title") or paper.get("title") or paper_id)
        row = [
            paper_id,
            title,
            _stringify(included),
            _stringify(paper.get("publisher")),
            _stringify(paper.get("published")),
            _stringify(paper.get("doi")),
            _stringify(link),
        ]

        for category_name in categories:
            payload = category_map.get(category_name) or {}
            row.append(_format_category_cell(payload))

        included_rank = 0 if included is True else 1
        sortable_rows.append((included_rank, title.lower(), row))

    sortable_rows.sort(key=lambda item: (item[0], item[1]))
    return [row for _, _, row in sortable_rows]


def _format_category_cell(payload: dict[str, Any]) -> str:
    paragraph = _stringify(payload.get("paragraph"))
    quotes = payload.get("quotes") or []
    if not isinstance(quotes, list):
        quotes = [quotes]
    cleaned_quotes = [str(item).strip() for item in quotes if str(item).strip() and str(item).strip() != "--"]

    parts: list[str] = []
    if paragraph:
        parts.append(f"Paragraph:\n{paragraph}")
    if cleaned_quotes:
        parts.append("Quotes:\n" + "\n".join(f"- {quote}" for quote in cleaned_quotes))
    return "\n\n".join(parts)


def _flatten_prisma(prisma: Any) -> list[list[str]]:
    if not isinstance(prisma, dict) or not prisma:
        return []

    rows: list[list[str]] = []
    identification = prisma.get("identification")
    if isinstance(identification, dict):
        for source, value in identification.items():
            rows.append([f"Identification: {source}", _stringify(value)])

    for key in [
        "after_dedup",
        "screened",
        "excluded_screening",
        "sought_retrieval",
        "not_retrieved",
        "assessed_eligibility",
        "excluded_eligibility",
        "included",
    ]:
        if key in prisma:
            rows.append([key.replace("_", " ").title(), _stringify(prisma.get(key))])
    return rows


def _build_criterion_specs(criteria: list[str]) -> list[dict[str, str]]:
    include_index = 0
    exclude_index = 0
    specs: list[dict[str, str]] = []

    for raw in criteria:
        criterion = str(raw).strip()
        if not criterion:
            continue

        upper = criterion.upper()
        if upper.startswith("EXCLUDE:"):
            exclude_index += 1
            specs.append(
                {
                    "code": f"EXCLUDE_{exclude_index:02d}",
                    "type": "Exclude",
                    "criterion": criterion,
                }
            )
        else:
            include_index += 1
            specs.append(
                {
                    "code": f"INCLUDE_{include_index:02d}",
                    "type": "Include",
                    "criterion": criterion,
                }
            )
    return specs


def _criterion_header_label(spec: dict[str, str], max_len: int = 48) -> str:
    code = str(spec.get("code") or "").strip()
    criterion = str(spec.get("criterion") or "").strip()
    label = _strip_criterion_prefix(criterion)
    if len(label) > max_len:
        label = label[: max_len - 3].rstrip() + "..."
    return f"{code}: {label}" if code else label


def _strip_criterion_prefix(text: str) -> str:
    stripped = text.strip()
    upper = stripped.upper()
    if upper.startswith("INCLUDE:"):
        return stripped[len("INCLUDE:"):].strip()
    if upper.startswith("EXCLUDE:"):
        return stripped[len("EXCLUDE:"):].strip()
    return stripped


def _normalize_screening_answer(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in {"YES", "NO", "INSUFFICIENT"}:
        return text
    if text in {"TRUE", "1"}:
        return "YES"
    if text in {"FALSE", "0"}:
        return "NO"
    return ""


def _write_section_title(ws, row: int, title: str) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = WRAP_TOP


def _write_table(
    ws,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    table_name: str,
) -> int:
    rows = list(rows)
    for offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP

    for row_index, row_values in enumerate(rows, start=start_row + 1):
        for col_index, value in enumerate(row_values, start=start_col):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = WRAP_TOP

    end_row = start_row + max(len(rows), 1)
    end_col = start_col + len(headers) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return end_row


def _write_headers(ws, row: int, start_col: int, headers: list[str]) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP


def _write_used_criteria_table(
    ws,
    start_row: int,
    criterion_specs: list[dict[str, str]],
) -> int:
    header_left = ws.cell(row=start_row, column=1, value="Column Name")
    header_left.fill = HEADER_FILL
    header_left.font = HEADER_FONT
    header_left.alignment = WRAP_TOP

    header_right = ws.cell(row=start_row, column=2, value="Used Criterion")
    header_right.fill = HEADER_FILL
    header_right.font = HEADER_FONT
    header_right.alignment = WRAP_TOP
    ws.cell(row=start_row, column=3).fill = HEADER_FILL
    ws.cell(row=start_row, column=3).font = HEADER_FONT
    ws.cell(row=start_row, column=3).alignment = WRAP_TOP
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)

    current_row = start_row + 1
    rows = criterion_specs or [{"code": "", "criterion": ""}]
    for idx, spec in enumerate(rows):
        code_cell = ws.cell(row=current_row, column=1, value=spec["code"])
        code_cell.alignment = WRAP_TOP
        criterion_cell = ws.cell(row=current_row, column=2, value=spec["criterion"])
        criterion_cell.alignment = WRAP_TOP
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        if idx % 2 == 0:
            code_cell.fill = SUBTLE_FILL
            criterion_cell.fill = SUBTLE_FILL
            ws.cell(row=current_row, column=3).fill = SUBTLE_FILL
        current_row += 1

    return current_row - 1


def _format_theme_text(categories: dict[str, Any]) -> str:
    lines: list[str] = []
    for name, description in categories.items():
        name_text = str(name).strip()
        desc_text = str(description).strip() if description is not None else ""
        if not name_text:
            continue
        if desc_text:
            lines.append(f"\"{name_text}\": \"{desc_text}\"")
        else:
            lines.append(name_text)
    return "\n".join(lines)


def _set_column_widths(ws, width_map: dict[str, float]) -> None:
    for column, width in width_map.items():
        ws.column_dimensions[column].width = width


def _set_row_heights(ws, start_row: int, end_row: int, height: float) -> None:
    for row_idx in range(start_row, end_row + 1):
        ws.row_dimensions[row_idx].height = height


def _join_lines(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item).strip() for item in value if str(item).strip())
    return _stringify(value)


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def main(argv: list[str] | None = None) -> int:
    argv = list(argv or sys.argv[1:])
    log_path = Path(argv[0]) if argv else DEFAULT_LOG_PATH
    output_path = Path(argv[1]) if len(argv) > 1 else DEFAULT_TEST_OUTPUT_PATH
    written = export_log_to_excel(log_path, output_path)
    print(str(written))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
